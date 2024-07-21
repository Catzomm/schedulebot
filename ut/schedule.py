import openpyxl
from win32com import client
import pythoncom
import os
from openpyxl.styles import Alignment
from ut.main import schedule_handler
from ut.convert_to_img import pdf_to_image


def schedule(teacher):
    try:
        workbook2 = schedule_handler.workbook4
        sheet2 = workbook2['schedule']
        k = teacher * 14
        for row in range(1, 13):
            for column in range(1, 14):
                value2 = schedule_handler.sheet3.cell(row=row + k, column=column).value
                cell2 = sheet2.cell(row=row, column=column)
                if value2 != None:
                    cell2.value = value2
                elif row >= 5:
                    cell2.value = ''

        workbook2.save('file/pdf/schedule.xlsx')

        script_dir = os.path.dirname(__file__)
        script_dir = '\\'.join(script_dir.split('\\')[:-1] + ['file\\pdf'])

        rel_path = "schedule.xlsx"
        rel_path2 = "Расписание.pdf"
        abs_file_path = os.path.join(script_dir, rel_path)
        abs_file_path2 = os.path.join(script_dir, rel_path2)
        pythoncom.CoInitialize()
        excel = client.DispatchEx("Excel.Application")
        sheets = excel.Workbooks.Open(abs_file_path, ReadOnly=True)
        work_sheets = sheets.Worksheets[0]
        work_sheets.PageSetup.Orientation = 2
        work_sheets.ExportAsFixedFormat(0, abs_file_path2)
        sheets.Close(False)
        excel.Quit()
        del excel
        pdf_file_path = "file/pdf/Расписание.pdf"
        output_image_path = "file/img/Расписание.png"  # Output image file path

        pdf_to_image(pdf_file_path, output_image_path)

    except Exception as e:
        print(repr(e))


def schedule2(class_id):
    try:
        k = 3 + class_id * 2
        workbook2 = openpyxl.load_workbook('file/pdf/schedule_sample.xlsx')
        sheet2 = workbook2['schedule']
        value = schedule_handler.sheet.cell(row=5, column=k).value
        cell2 = sheet2.cell(row=1, column=1)
        cell2.value = str(value) + ' класс'
        h = -48
        for column in range(3, 10, 6):
            h += 48
            for row in range(6, 53, 2):
                if type(schedule_handler.sheet.cell(row=row + 2 + h, column=k)).__name__ == 'MergedCell':
                    sheet2.merge_cells(start_row=row - 1, start_column=column, end_row=row, end_column=column)
                    sheet2.cell(row=row - 1, column=column).alignment = Alignment(horizontal="center",
                                                                                  vertical="center")
                    value = schedule_handler.sheet.cell(row=row + 1 + h, column=k).value
                    cell2 = sheet2.cell(row=row-1, column=column)

                    sheet2.merge_cells(start_row=row - 1, start_column=column + 1, end_row=row, end_column=column + 1)
                    sheet2.cell(row=row - 1, column=column + 1).alignment = Alignment(horizontal="center",
                                                                                      vertical="center")
                    value3 = schedule_handler.sheet.cell(row=row + 1 + h, column=k + 1).value
                    cell3 = sheet2.cell(row=row - 1, column=column + 1)
                    if value != None:
                        cell2.value = value
                        cell3.value = value3
                else:
                    value = schedule_handler.sheet.cell(row=row + 1 + h, column=k).value
                    cell2 = sheet2.cell(row=row - 1, column=column)
                    value3 = schedule_handler.sheet.cell(row=row + 1 + h, column=k + 1).value
                    cell3 = sheet2.cell(row=row - 1, column=column + 1)
                    if value != None:
                        cell2.value = value
                        cell3.value = value3

                    value = schedule_handler.sheet.cell(row=row + 2 + h, column=k).value
                    cell2 = sheet2.cell(row=row, column=column)
                    value3 = schedule_handler.sheet.cell(row=row + 2 + h, column=k + 1).value
                    cell3 = sheet2.cell(row=row, column=column + 1)
                    if value != None:
                        cell2.value = value
                        cell3.value = value3

        workbook2.save('file/pdf/schedule2.xlsx')
        workbook2.close()

        script_dir = os.path.dirname(__file__)
        script_dir = '\\'.join(script_dir.split('\\')[:-1] + ['file\\pdf'])

        rel_path = "schedule2.xlsx"
        rel_path2 = "Расписание уроков.pdf"
        abs_file_path = os.path.join(script_dir, rel_path)
        abs_file_path2 = os.path.join(script_dir, rel_path2)
        pythoncom.CoInitialize()
        excel = client.DispatchEx("Excel.Application")
        sheets = excel.Workbooks.Open(abs_file_path, ReadOnly=True)
        work_sheets = sheets.Worksheets[0]
        work_sheets.ExportAsFixedFormat(0, abs_file_path2)
        sheets.Close(False)
        excel.Quit()
        del excel
        pdf_file_path = "file/pdf/Расписание уроков.pdf"
        output_image_path = "file/img/Расписание уроков.png"  # Output image file path

        pdf_to_image(pdf_file_path, output_image_path)

    except Exception as e:
        print(repr(e))