import openpyxl
from ut.convert_to_img import xlsx_to_img

class ScheduleHandler:
    def __init__(self):
        self.new_of_lessons = {}
        self.old_of_lessons = {}
        self.old_lessons = []
        self.new_lessons = []
        self.last_class = 8

    def new_calls(self):
        self.time3 = []
        for row in range(5, 20, 2):
            if type(self.sheet2.cell(row=row + 1, column=3)).__name__ == 'MergedCell':
                if self.sheet2.cell(row=row , column=3).value == None:
                    print(self.file)
                    if self.file == 'Суббота':
                        self.time3 = self.time2
                        break
                    elif self.file == 'Понедельник':
                        self.time3 = self.time0
                        break
                    else:
                        self.time3 = self.time
                        break
                else:
                    time1 = self.sheet2.cell(row=row, column=3).value
                    if " " in time1:
                        time1 = time1.replace(" ", "")
                    if "-" in time1:
                        time1 = time1.replace("-", " - ")
                    self.time3 += [f'{row // 2 - 1}.  {time1}']
            elif self.sheet2.cell(row=row + 1, column=3).value == None and self.sheet2.cell(row=row, column=3).value == None:
                if self.file == 'Суббота':
                    self.time3 = self.time2
                    break
                elif self.file == 'Понедельник':
                    self.time3 = self.time0
                    break
                else:
                    self.time3 = self.time
                    break
            elif self.sheet2.cell(row=row, column=3).value == None:
                time1 = self.sheet2.cell(row=row + 1, column=3).value
                if " " in time1:
                    time1 = time1.replace(" ", "")
                if "-" in time1:
                    time1 = time1.replace("-", " - ")
                self.time3 += [f'{row // 2 - 1}.  {time1}']
            elif self.sheet2.cell(row=row + 1, column=3).value == None:
                time1 = self.sheet2.cell(row=row, column=3).value
                if " " in time1:
                    time1 = time1.replace(" ", "")
                if "-" in time1:
                    time1 = time1.replace("-", " - ")
                self.time3 += [f'{row // 2 - 1}.  {time1}']
            else:
                time1 = self.sheet2.cell(row=row, column=3).value
                time2 = self.sheet2.cell(row=row + 1, column=3).value
                if " " in time2:
                    time2 = time2.replace(" ", "")
                if " " in time1:
                    time1 = time1.replace(" ", "")
                self.time3 += [f'{row // 2 - 1}.  {time1} - {time2}']

    def calls(self):
        self.time0 = []
        self.time = []
        self.time2 = []
        for row in range(2, 17, 2):
            time0 = self.call_schedule.cell(row=row, column=2).value.replace(" ", "").replace("-", " - ")
            time = self.call_schedule.cell(row=row, column=5).value.replace(" ", "").replace("-", " - ")
            time2 = self.call_schedule.cell(row=row, column=8).value.replace(" ", "").replace("-", " - ")
            self.time0 += [f'{row // 2}.  {time0}']
            self.time += [f'{row// 2}.  {time}']
            self.time2 += [f'{row // 2}.  {time2}']

    def main_schedule(self):
        workbook = openpyxl.load_workbook('Расписание.xlsx')
        names = workbook.sheetnames

        for name in names:
            if '5-11' in name.strip():
                self.sheet = workbook[name]
                break

        for name in names:
            if 'звонки' in name.strip().lower():
                self.call_schedule = workbook[name]
                break

        for name in names:
            if 'учителя' == name.strip().lower():
                self.sheet3 = workbook[name]
                break

        self.calls()
        workbook.close()

    def teacher_schedule(self):
        self.workbook4 = openpyxl.load_workbook('file/pdf/schedule.xlsx')
        self.sheet4 = self.workbook4['schedule']
        self.workbook5 = openpyxl.load_workbook('file/pdf/schedule_sample.xlsx')
        self.sheet5 = self.workbook4['schedule']

    def lessons_schedule(self, row, column, row2, column2):
        lesson = self.sheet.cell(row=row, column=column).value
        lesson2 = self.sheet2.cell(row=row2, column=column2).value
        lesson3 = self.sheet.cell(row=row + 1, column=column).value
        lesson4 = self.sheet2.cell(row=row2 + 1, column=column2).value

        if lesson != None:
            lesson = lesson.strip().lower()

        if lesson2 != None:
            lesson2 = lesson2.strip().lower()

        if lesson3 != None:
            lesson3 = lesson3.strip().lower()

        if lesson4 != None:
            lesson4 = lesson4.strip().lower()

        return lesson, lesson2, lesson3, lesson4

    def schedule(self):
        columns = self.sheet.max_column

        k = 5
        k2 = 3
        column2 = 2
        new_lessons = []
        old_lessons = []

        days = {'Понедельник': 0,
                'Вторник': 1,
                'Среда': 2,
                'Четверг': 3,
                'Пятница': 4,
                'Суббота': 5}

        rows = days[self.file]

        flag = False
        flag_2 = True
        while True:
            if self.sheet.cell(row=k, column=3).value == None:
                k += 1
            elif not self.sheet.cell(row=k, column=3).value.strip()[:-1].isdigit():
                k += 1
            else:
                break

        self.new_calls()

        # Анализ изменения в расписании
        for i in range(0, columns, 2):
            event = ''
            column = i + 3
            if not self.sheet.cell(row=k, column=column).value:
                break
            if not self.sheet.cell(row=k, column=column).value.strip()[:-1].isdigit():
                break
            column2 += 2
            if not self.sheet2.cell(row=k2, column=column2).value:
                if flag:
                    if flag_2:
                        flag_2 = False
                        self.last_class = int(self.sheet2.cell(row=k2, column=column2-2).value[:-1])
                    flag = False
                    k2 += 21
                else:
                    flag = True
                    k2 += 19
                column2 = 4

            for j in range(2 + 16 * rows, 18 + 16 * rows, 2):
                row = k + j
                row2 = k2 + j - 16 * rows

                #lesson, lesson2, lesson3, lesson4 = self.lessons_schedule(row, column, row2, column2)
                lesson = self.sheet.cell(row=row, column=column).value
                lesson2 = self.sheet2.cell(row=row2, column=column2).value
                lesson3 = self.sheet.cell(row=row + 1, column=column).value
                lesson4 = self.sheet2.cell(row=row2 + 1, column=column2).value

                classroom2 = self.sheet2.cell(row=row2, column=column2 + 1).value
                classroom4 = self.sheet2.cell(row=row2 + 1, column=column2 + 1).value
                number = (j - 16 * rows) // 2

                if type(self.sheet2.cell(row=row2, column=column2)).__name__ == 'MergedCell':
                    if lesson:
                        old_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson])
                    if event:
                        new_lessons.append([self.sheet.cell(row=k, column=column).value, number, event, classroom2])
                elif type(self.sheet.cell(row=row + 1, column=column)).__name__ == 'MergedCell':
                    event = lesson2
                    if type(self.sheet2.cell(row=row2 + 1, column=column2)).__name__ == 'MergedCell':
                        if lesson != lesson2 or classroom2 in ['задание', 'вебинар']: #объединение двух if
                            if lesson:
                                old_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson])
                            if lesson2:
                                new_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson2, classroom2])
                    else:
                        if lesson:
                            old_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson])
                        if lesson2:
                            new_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson2, 'I', classroom2])
                        if lesson4:
                            new_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson4, 'II', classroom4])

                else:
                    event = lesson2
                    if type(self.sheet2.cell(row=row2 + 1, column=column2)).__name__ == 'MergedCell' and not (
                            type(self.sheet.cell(row=row + 1, column=column)).__name__ == 'MergedCell'):
                        if lesson:
                            old_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson, 'I'])
                        if lesson2:
                            new_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson2, classroom2])
                        if lesson3:
                            old_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson3, 'II'])

                    else:
                        if lesson != lesson2 or classroom2 in ['задание', 'вебинар']: #Объединение двух if ###################################
                            if lesson:
                                old_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson, 'I'])
                            if lesson2:
                                new_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson2, 'I', classroom2])
                        if lesson3 != lesson4 or classroom4 in ['задание', 'вебинар']: #Объединение двух if #####################
                            if lesson3:
                                old_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson3, 'II'])
                            if lesson4:
                                new_lessons.append([self.sheet.cell(row=k, column=column).value, number, lesson4, 'II', classroom4])
        self.old_lessons = old_lessons
        self.new_lessons = new_lessons


    def old(self):
        old_teachers = {}


        days = {'Понедельник': 2,
                'Вторник': 4,
                'Среда': 6,
                'Четверг': 8,
                'Пятница': 10,
                'Суббота': 12}

        column = days[self.file]
        rows = self.sheet3.max_row + 5
        if column != 12:
            for lesson_info in self.old_lessons:
                flag = True
                for i in range(rows // 14):
                    row = i * 14 + lesson_info[1] + 4
                    lesson = self.sheet3.cell(row=row, column=column).value
                    if len(lesson_info) == 3:
                        new_lesson = '-'.join([lesson_info[0], lesson_info[2]])
                    else:
                        new_lesson = '-'.join([lesson_info[0], lesson_info[2], lesson_info[3]])
                    if lesson != None:
                        if lesson.strip().lower() == new_lesson.lower():
                            flag = False
                            teacher = self.sheet3.cell(row=i * 14 + 1, column=1).value
                            if teacher in old_teachers:
                                a = old_teachers[teacher]
                                a.append(lesson_info)
                                old_teachers[teacher] = a
                            else:
                                old_teachers[teacher] = [lesson_info]
                            break
                if flag:
                    print(new_lesson)
        else:
            for lesson_info in self.old_lessons:
                for i in range(rows // 14):
                    row = i * 14 + lesson_info[1] + 4
                    lesson = self.sheet3.cell(row=row, column=column).value
                    lesson2 = []
                    if ',' in lesson:
                        for j in lesson.strip().split(','):
                            if '-' in j:
                                a = j.split('-')[0]
                            else:
                                a = j
                            if a.strip()[:-1].isdigit():
                                if '-' in lesson:
                                    lesson2.append(a.strip() + lesson[lesson.index('-'):])
                    else:
                        lesson2.append(lesson)

                    if len(lesson_info) == 3:
                        new_lesson = '-'.join([lesson_info[0], lesson_info[2]])
                    else:
                        new_lesson = '-'.join([lesson_info[0], lesson_info[2], lesson_info[3]])

                    if new_lesson in lesson2:
                        teacher = self.sheet3.cell(row=i * 14 + 1, column=1).value
                        if teacher in old_teachers:
                            a = old_teachers[teacher]
                            a.append(lesson_info)
                            old_teachers[teacher] = a
                        else:
                            old_teachers[teacher] = [lesson_info]
                        break



        self.old_teachers = old_teachers


    def new(self):
        rows = self.sheet3.max_row + 5
        new_teachers = {}

        columnss = [2, 4, 6, 8, 10, 12]
        days = {'Понедельник': 2,
                'Вторник': 4,
                'Среда': 6,
                'Четверг': 8,
                'Пятница': 10,
                'Суббота': 12}

        del columnss[days[self.file] // 2 - 1]
        columnss.insert(0, days[self.file])

        for lesson_info in self.new_lessons:
            teacherss = []
            flag = False

            for column in columnss:
                for j in range(rows // 14 * 8):
                    row = j % 8 + j // 8 * 14 + 5
                    lesson = self.sheet3.cell(row=row, column=column).value
                    lesson2 = []

                    if lesson != None:
                        if ',' in lesson:
                            for j in lesson.strip().split(','):
                                # print(i)
                                if '-' in j:
                                    a = j.split('-')[0]
                                else:
                                    a = j
                                if a.strip()[:-1].isdigit():
                                    if '-' in lesson:
                                        lesson2.append(a.strip() + lesson[lesson.index('-'):])
                        else:
                            lesson2.append(lesson.strip())

                        if len(lesson_info) == 4:
                            new_lesson = '-'.join([lesson_info[0], lesson_info[2]])
                        else:
                            new_lesson = '-'.join([lesson_info[0], lesson_info[2], lesson_info[3]])

                        if new_lesson.strip() in lesson2:
                            teacher = self.sheet3.cell(row=row - row % 14 + 1, column=1).value
                            if teacher in new_teachers:
                                a = new_teachers[teacher]
                                a.append(lesson_info)
                                new_teachers[teacher] = a
                                flag = True
                                teacherss = []
                            else:
                                new_teachers[teacher] = [lesson_info]
                                flag = True
                                teacherss = []
                        elif len(new_lesson.split('-')) == 2 and len(lesson.split('-')) == 3:
                            if '-'.join(lesson.split('-')) == new_lesson:
                                teacher = self.sheet3.cell(row=row - row % 14 + 1, column=1).value
                                if not (teacher in teacherss):
                                    teacherss.append(teacher)
                        if flag:
                            break
                if flag:
                    break
            if teacherss:
                for teacher in teacherss:
                    if teacher in new_teachers:
                        a = new_teachers[teacher]
                        a.append(lesson_info)
                        new_teachers[teacher] = a
                    else:
                        new_teachers[teacher] = [lesson_info]
        self.new_teachers = new_teachers


    def main(self, file):
        self.file = file
        workbook2 = openpyxl.load_workbook(f'{file}.xlsx')
        if '5-11' in workbook2.sheetnames:
            self.sheet2 = workbook2['5-11']
        else:
            self.sheet2 = workbook2[workbook2.sheetnames[0]]

        # Создание картинки

        xlsx_to_img(file)

        # Работа с расписанием

        self.schedule()
        print(self.old_lessons)
        print(self.new_lessons)
        if file == 'Суббота':
            if self.time3 == self.time2:
                self.new_time = []
            else:
                self.new_time = self.time3
        elif file == 'Понедельник':
            if self.time3 == self.time0:
                self.new_time = []
            else:
                self.new_time = self.time3
        else:
            if self.time3 == self.time:
                self.new_time = []
            else:
                self.new_time = self.time3

        self.new_of_lessons = {}
        self.old_of_lessons = {}
        for lesson in self.old_lessons:
            if lesson[0] in self.old_of_lessons:
                lessons = self.old_of_lessons[lesson[0]]
                lessons.append(lesson)
                self.old_of_lessons[lesson[0]] = lessons
            else:
                self.old_of_lessons[lesson[0]] = [lesson]

        for lesson in self.new_lessons:
            if lesson[0] in self.new_of_lessons:
                lessons = self.new_of_lessons[lesson[0]]
                lessons.append(lesson)
                self.new_of_lessons[lesson[0]] = lessons
            else:
                self.new_of_lessons[lesson[0]] = [lesson]

        self.old()
        self.new()



schedule_handler = ScheduleHandler()
