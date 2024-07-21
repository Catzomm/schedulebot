import openpyxl
from ut.main import schedule_handler

help_message = 'Вас приветствует Telegram bot \n<b>"КФМЛ Расписание"</b>\n\n' \
               'Для получения подробной информации о возможностях бота, выберите соответствующий пункт'

help_message_1 = '<b>Регистрация</b>\n\n' \
                 '- При первом взаимодействии с ботом команда /start вводится автоматически \n' \
                 '- Выберите, кто Вы, учитель или ученик\n' \
                 '- Если Вы учитель, то найдите себя в списке, а если Вы ученик, то выберите класс\n' \
                 '- Бот запрашивает у пользователя информацию о нём\n' \
                 '- Бот сохраняет информацию о пользователе и отправляет подтверждение регистрации'

help_message_2 = '<b>Отправка изменений в расписании</b>\n\n' \
                 '- Бот периодически проверяет наличие изменений в расписании\n' \
                 '- При обнаружении изменений бот отправляет зарегистрированным пользователям сообщение с изменением в расписании в формате:\n\n' \
                 '№ урока, урок в текущем расписании, класс, подгруппа (если есть) ➡️ \nурок в измененном расписании, класс, подгруппа (если есть)' \
                 '\n\nНапример:\n2\ufe0f\u20e3 Англ. язык 10Б I \u27a1\ufe0f Математика 10Б I'

help_message_3 = '<b>Получение текущего расписания на неделю</b>\n\n' \
                 '- Отправьте сообщение "Мое расписание" или выберите соответствующий пункт меню, чтобы запросить актуальное расписание на неделю\n' \
                 '- Бот отправляет расписание пользователю в виде картинки'

help_message_4 = '<b>Получение текущего расписания звонков на неделю</b>\n\n' \
                 '- Отправьте сообщение "Расписание звонков" или выберите соответствующий пункт меню, чтобы запросить актуальное расписание звонков на неделю\n' \
                 '- Бот отправляет расписание звонков пользователю в виде сообщения'


def list_teachers():
    list_teacher = []
    rows = (schedule_handler.sheet3.max_row + 5) // 14

    for i in range(rows):
        teacher = schedule_handler.sheet3.cell(row=i * 14 + 1, column=1).value
        list_teacher.append(teacher)
    return list_teacher


def list_class():
    list_class_1 = []
    columns = (schedule_handler.sheet.max_column - 2) // 2

    for i in range(columns):
        number_class = schedule_handler.sheet.cell(row=5, column=i * 2 + 3).value
        if number_class == None or number_class.isalpha() or not((str(number_class)[0]).isdigit()):
            break
        else:
            list_class_1.append(number_class)

    return list_class_1


def write_telegram(teacher, id, file):
    flag = True
    workbook2 = openpyxl.load_workbook('file/teachers_telegram.xlsx')
    sheet2 = workbook2.active

    workbook3 = openpyxl.load_workbook('file/class_telegram.xlsx')
    sheet3 = workbook3.active

    for i in range(1, sheet2.max_row + 1):
        if sheet2.cell(row=i, column=2).value != None:
            value_id = sheet2.cell(row=i, column=2).value.split(', ')
            if str(id) in value_id:
                cell = sheet2.cell(row=i, column=2)
                value_id.remove(str(id))
                cell.value = ', '.join(value_id)
                break

    for i in range(1, sheet3.max_row + 1):
        if sheet3.cell(row=i, column=2).value != None:
            value_id = sheet3.cell(row=i, column=2).value.split(', ')
            if str(id) in value_id:
                cell = sheet3.cell(row=i, column=2)
                value_id.remove(str(id))
                cell.value = ', '.join(value_id)
                break

    workbook2.save('file/teachers_telegram.xlsx')
    workbook3.save('file/class_telegram.xlsx')

    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for i in range(1, sheet.max_row + 1):
        value = sheet.cell(row=i, column=1).value
        if value == teacher:
            flag = False
            value_id = str(sheet.cell(row=i, column=2).value)
            if value_id == 'None':
                value_id = str(id)
            else:
                value_id += ', ' + str(id)
            cell = sheet.cell(row=i, column=2)
            cell.value = value_id
            break

    if flag:
        cell = sheet.cell(row=i + 1, column=1)
        cell.value = teacher
        cell = sheet.cell(row=i + 1, column=2)
        cell.value = str(id)

    workbook.save(file)
    read_teachers(file)


def read_teachers(file):
    teachers = {}
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for i in range(1, sheet.max_row + 1):
        value = sheet.cell(row=i, column=1).value
        value2 = sheet.cell(row=i, column=2).value
        if value2 == None:
            value_id = []
        else:
            value_id = str(value2).split(', ')
        teachers[value] = value_id

    new_teachers =[]
    for element in teachers.keys():
        new_teachers += teachers[element]

    if file == 'file/teachers_telegram.xlsx':
        schedule_handler.teacher_id = new_teachers
    elif file == 'file/class_telegram.xlsx':
        schedule_handler.class_id = new_teachers
    return teachers


def teacher_id(id):
    read_teacher = read_teachers('file/teachers_telegram.xlsx')
    list_teacher = list_teachers()

    for teacher in read_teacher:
        if str(id) in read_teacher[teacher]:
            return list_teacher.index(teacher)


def class_id(id):
    read_class = read_teachers('file/class_telegram.xlsx')
    list_class_id = list_class()

    for name in read_class:
        if str(id) in read_class[name]:
            return list_class_id.index(name)


def beautiful_schedule(old_lesson, new_lesson):
    emoji = ['1⃣', '2⃣', '3⃣', '4⃣', '5⃣', '6⃣', '7⃣', '8⃣']
    mess = ""
    for number in range(1, 9):
        if number in new_lesson and number in old_lesson:
            if type(old_lesson[number][0]) == list and type(new_lesson[number][0]) == list:
                mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number][0])}  ➡️  {" ".join(new_lesson[number][0])}\n\n'
                mess += f'         {" ".join(old_lesson[number][1])}  ➡️  {" ".join(new_lesson[number][1])}\n\n'
            elif type(old_lesson[number][0]) == list:
                if len(new_lesson[number]) == 2:
                    mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number][0])}  ➡️  {" ".join(new_lesson[number])}\n\n'
                    mess += f'         {" ".join(old_lesson[number][1])}  ➡️  {" ".join(new_lesson[number])}\n\n'
                elif new_lesson[number][2] == 'I':
                    mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number][0])}  ➡️  {" ".join(new_lesson[number])}\n\n'
                    mess += f'         {" ".join(old_lesson[number][1])}  ➡️  Нет урока\n\n'
                else:
                    mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number][0])}  ➡️  Нет урока\n\n'
                    mess += f'         {" ".join(old_lesson[number][1])}  ➡️  {" ".join(new_lesson[number])}\n\n'
            elif type(new_lesson[number][0]) == list:
                if len(new_lesson[number]) == 2:
                    mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number])}  ➡️  {" ".join(new_lesson[number][0])}\n\n'
                    mess += f'         {" ".join(old_lesson[number])}  ➡️  {" ".join(new_lesson[number][1])}\n\n'
                elif old_lesson[number][2] == 'I':
                    mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number])}  ➡️  {" ".join(new_lesson[number][0])}\n\n'
                    mess += f'         Нет урока  ➡️  {" ".join(new_lesson[number][1])}\n\n'
                else:
                    mess += f'{emoji[number - 1]}  Нет урока  ➡️  {" ".join(new_lesson[number][0])}\n\n'
                    mess += f'         {" ".join(old_lesson[number][1])}  ➡️  {" ".join(new_lesson[number][1])}\n\n'
            else:
                mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number])}  ➡️  {" ".join(new_lesson[number])}\n\n'
        elif number in new_lesson:
            if type(new_lesson[number][0]) == list:
                mess += f'{emoji[number - 1]}  Нет урока  ➡️  {" ".join(new_lesson[number][0])}\n\n'
                mess += f'       Нет урока  ➡️  {" ".join(new_lesson[number][1])}\n\n'
            else:
                mess += f'{emoji[number - 1]}  Нет урока  ➡️  {" ".join(new_lesson[number])}\n\n'
        elif number in old_lesson:
            if type(old_lesson[number][0]) == list:
                mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number][0])}  ➡️  Нет урока\n\n'
                mess += f'       {" ".join(old_lesson[number][0])}  ➡️  Нет урока\n\n'
            else:
                mess += f'{emoji[number - 1]}  {" ".join(old_lesson[number])}  ➡️  Нет урока\n\n'
    return mess


def sending_old(name, users):
    new_lesson = {}

    for lesson in users[name]:
        if lesson[1] in new_lesson and name[:-1].isdigit(): #ученики
            lesson2 = new_lesson[lesson[1]]
            new_lesson[lesson[1]] = [lesson2, [lesson[2], lesson[0], lesson[3]]]

        elif lesson[1] in new_lesson:
            if len(lesson) == 4:
                if len(new_lesson[lesson[1]]) == 3:
                    lesson2 = new_lesson[lesson[1]]
                    lesson2[1] = f'{lesson2[1]} ({lesson2[2]}), {lesson[0]} ({lesson[3]})'
                    del lesson2[2]
                    new_lesson[lesson[1]] = lesson2
                else:
                    lesson2 = new_lesson[lesson[1]]
                    lesson2[1] = f'{lesson2[1]}, {lesson[0]} ({lesson[3]})'
                    new_lesson[lesson[1]] = lesson2

            else:
                if len(new_lesson[lesson[1]]) == 3:
                    lesson2 = new_lesson[lesson[1]]
                    lesson2[1] = lesson2[1] + f'({lesson2[2]})' + ', ' + lesson[0]
                    del lesson2[2]
                    new_lesson[lesson[1]] = lesson2
                else:
                    lesson2 = new_lesson[lesson[1]]
                    lesson2[1] = lesson2[1] + ', ' + lesson[0]
                    new_lesson[lesson[1]] = lesson2

        else:
            if len(lesson) == 4:
                new_lesson[lesson[1]] = [lesson[2], lesson[0], lesson[3]]
            else:
                new_lesson[lesson[1]] = [lesson[2], lesson[0]]

    return new_lesson


def sending_new(name, users):
    new_lesson = {}

    for lesson in users[name]:
        if name[:-1].isdigit():
            if lesson[1] in new_lesson: #ученики
                lesson2 = new_lesson[lesson[1]]
                if len(lesson) == 5 and lesson[-1] in ['задание', 'вебинар']:
                    new_lesson[lesson[1]] = [lesson2, [lesson[2], lesson[0], f'{lesson[3]} ({lesson[4]})']]
                else:
                    new_lesson[lesson[1]] = [lesson2, [lesson[2], lesson[0], lesson[3]]]
            else:
                if len(lesson) == 5 and lesson[-1] in ['задание', 'вебинар']:
                    new_lesson[lesson[1]] = [lesson[2], lesson[0], f'{lesson[3]} ({lesson[4]})']
                elif len(lesson) == 5:
                    new_lesson[lesson[1]] = [lesson[2], lesson[0], lesson[3]]
                else:
                    if lesson[-1] in ['задание', 'вебинар']:
                        new_lesson[lesson[1]] = [lesson[2], f'{lesson[0]} ({lesson[3]})']
                    else:
                        new_lesson[lesson[1]] = [lesson[2], lesson[0]]

        elif lesson[1] in new_lesson:
            if len(lesson) == 5:
                if len(new_lesson[lesson[1]]) == 3:
                    lesson2 = new_lesson[lesson[1]]
                    lesson2[1] = f'{lesson2[1]} ({lesson2[2]}), {lesson[0]} ({lesson[3]})'
                    del lesson2[2]
                    new_lesson[lesson[1]] = lesson2
                else:
                    lesson2 = new_lesson[lesson[1]]
                    lesson2[1] = f'{lesson2[1]}, {lesson[0]} ({lesson[3]})'
                    new_lesson[lesson[1]] = lesson2

            else:
                if len(new_lesson[lesson[1]]) == 3:
                    lesson2 = new_lesson[lesson[1]]
                    lesson2[1] = lesson2[1] + f'({lesson2[2]})' + ', ' + lesson[0]
                    del lesson2[2]
                    new_lesson[lesson[1]] = lesson2
                else:
                    lesson2 = new_lesson[lesson[1]]
                    lesson2[1] = lesson2[1] + ', ' + lesson[0]
                    new_lesson[lesson[1]] = lesson2

        else:
            if len(lesson) == 5 and lesson[-1] in ['задание', 'вебинар']:
                new_lesson[lesson[1]] = [lesson[2], lesson[0], f'{lesson[3]} ({lesson[4]})']
            elif len(lesson) == 5:
                new_lesson[lesson[1]] = [lesson[2], lesson[0], lesson[3]]
            else:
                if lesson[-1] in ['задание', 'вебинар']:
                    new_lesson[lesson[1]] = [lesson[2], f'{lesson[0]} ({lesson[3]})']
                else:
                    new_lesson[lesson[1]] = [lesson[2], lesson[0]]

    return new_lesson